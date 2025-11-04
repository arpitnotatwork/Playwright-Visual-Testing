import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def write_results_to_excel(results, file_path=None):
    # If no file_path provided, create timestamped default
    if not file_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        reports_dir = os.path.join(os.getcwd(), "reports")
        os.makedirs(reports_dir, exist_ok=True)
        file_path = os.path.join(reports_dir, f"broken_links_report_{timestamp}.xlsx")
    else:
        # If a path is provided, add timestamp to avoid overwriting
        base, ext = os.path.splitext(file_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = f"{base}_{timestamp}{ext}"
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Broken Links Report"

    headers = ["Page URL", "Link", "HTTP Code", "Meaning", "Result"]
    ws.append(headers)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in results:
        ws.append(row)
        if row[-1] == "Fail":
            for cell in ws[ws.max_row]:
                cell.fill = red_fill

    wb.save(file_path)
    return file_path
